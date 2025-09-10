import os
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ====== CONFIG ======
TEMPLATE_FILE = r"template.xlsx"          # <- Your template workbook (single-sheet recommended)
ROOT_DIR = r"path\to\your\root\directory" # <- Root folder containing subfolders with Excel files
TARGET_SHEET_TITLE = "CUI Cover Page"
CUIStr = "CUI"  # substring to detect existing CUI sheets
# ====================


def load_template_sheet(path: str | Path) -> Worksheet:
    """Load the template workbook and return its (first) sheet."""
    twb = load_workbook(path)
    if not twb.sheetnames:
        raise ValueError("Template workbook has no sheets.")
    return twb[twb.sheetnames[0]]


def find_cui_sheet(wb) -> Optional[Worksheet]:
    """Return the first worksheet whose name contains 'CUI' (case-sensitive by default)."""
    for ws in wb.worksheets:
        if CUIStr in ws.title:
            return ws
    return None


def move_leftmost(wb, ws: Worksheet) -> None:
    """
    Move the given worksheet to index 0 (leftmost).
    Prefer the public API move_sheet; fallback to _sheets manipulation if needed.
    """
    try:
        idx = wb.worksheets.index(ws)
        if idx != 0:
            wb.move_sheet(ws, -idx)
    except Exception:
        # Fallback: manipulate private list (used commonly in practice)
        sheets = wb._sheets  # type: ignore[attr-defined]
        i = sheets.index(ws)
        if i != 0:
            sheets.insert(0, sheets.pop(i))


def clone_sheet_into_workbook(template_ws: Worksheet, wb, title: str) -> Worksheet:
    """
    Create a new sheet in wb and copy values + common formatting from template_ws.
    Returns the new sheet.
    """
    # Create new sheet at the end; we will move it leftmost later.
    new_ws = wb.create_sheet(title=title)

    # Copy sheet-level properties (basic)
    new_ws.sheet_format.defaultColWidth = template_ws.sheet_format.defaultColWidth
    new_ws.sheet_format.defaultRowHeight = template_ws.sheet_format.defaultRowHeight
    new_ws.page_setup = template_ws.page_setup
    new_ws.page_margins = template_ws.page_margins
    new_ws.print_options = template_ws.print_options

    # Copy column dimensions (widths)
    for col_letter, dim in template_ws.column_dimensions.items():
        new_dim = new_ws.column_dimensions[col_letter]
        new_dim.width = dim.width
        new_dim.bestFit = dim.bestFit
        new_dim.customWidth = dim.customWidth

    # Copy row dimensions (heights)
    for row_idx, dim in template_ws.row_dimensions.items():
        new_dim = new_ws.row_dimensions[row_idx]
        new_dim.height = dim.height

    # Copy merged cells
    if template_ws.merged_cells.ranges:
        for mrange in list(template_ws.merged_cells.ranges):
            new_ws.merge_cells(str(mrange))

    # Copy cell values and styles
    for row in template_ws.iter_rows():
        for cell in row:
            tgt = new_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            # Styles (guard against None)
            if cell.has_style:
                tgt.font = cell.font
                tgt.fill = cell.fill
                tgt.border = cell.border
                tgt.alignment = cell.alignment
                tgt.number_format = cell.number_format
                tgt.protection = cell.protection

    # Copy freeze panes
    new_ws.freeze_panes = template_ws.freeze_panes

    return new_ws


def process_file(file_path: Path, template_ws: Worksheet) -> None:
    """Ensure a CUI sheet exists and is leftmost; otherwise insert from template and move leftmost."""
    try:
        keep_vba = file_path.suffix.lower() == ".xlsm"
        wb = load_workbook(file_path, keep_vba=keep_vba)
    except Exception as e:
        print(f"❌ Could not open {file_path}: {e}")
        return

    # If any CUI-like sheet exists, move it leftmost
    cui_ws = find_cui_sheet(wb)
    if cui_ws is not None:
        move_leftmost(wb, cui_ws)
        try:
            wb.save(file_path)
            print(f"↔️ Moved existing '{cui_ws.title}' to leftmost in {file_path}")
        except Exception as e:
            print(f"❌ Could not save {file_path}: {e}")
        return

    # Otherwise: insert cloned template named "CUI Cover Page", then move leftmost
    # If a sheet already exists with exactly that title, append a suffix to avoid conflicts, then rename after move.
    insert_title = TARGET_SHEET_TITLE
    counter = 1
    while insert_title in wb.sheetnames:
        counter += 1
        insert_title = f"{TARGET_SHEET_TITLE} ({counter})"

    new_ws = clone_sheet_into_workbook(template_ws, wb, insert_title)
    # Rename to canonical title if we had to use a temporary unique name (remove duplicates if needed)
    if insert_title != TARGET_SHEET_TITLE:
        # If the canonical name exists, delete it (we are replacing)
        if TARGET_SHEET_TITLE in wb.sheetnames:
            del wb[TARGET_SHEET_TITLE]
        new_ws.title = TARGET_SHEET_TITLE

    move_leftmost(wb, new_ws)

    try:
        wb.save(file_path)
        print(f"➕ Inserted '{TARGET_SHEET_TITLE}' (leftmost) into {file_path}")
    except Exception as e:
        print(f"❌ Could not save {file_path}: {e}")


def main():
    template_ws = load_template_sheet(TEMPLATE_FILE)

    for root, _, files in os.walk(ROOT_DIR):
        for name in files:
            if name.startswith("~$"):  # skip Excel lock files
                continue
            if not name.lower().endswith((".xlsx", ".xlsm")):
                continue
            file_path = Path(root) / name
            process_file(file_path, template_ws)


if __name__ == "__main__":
    main()